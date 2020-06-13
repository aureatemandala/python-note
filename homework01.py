#-*- encoding=utf-8 -*-

from openpyxl import load_workbook  #读取excel文件
from openpyxl import Workbook       #创建excel文件

workbook = load_workbook(filename='./files/test.xlsx',data_only=True)
sheet = workbook['Sheet3']

tmp_workbook = Workbook()
tmp_sheet = tmp_workbook.active
tmp_sheet.title = '书单'

keyword = ''
for item in sheet[1]:
    if item.value == '国籍':
        keyword = item.column -1

print(keyword)

for row in sheet.rows:
    cell = row[keyword]
    if cell.value == '英国':
        data_list = []
        for cells in row:
            data_list.append(cells.value)
        tmp_sheet.append(data_list)
            



tmp_workbook.save(filename='./files/homework.xlsx')
