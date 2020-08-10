#-*- coding=utf-8 -*-


from time import sleep
from openpyxl import load_workbook          
from bs4 import BeautifulSoup
import re
import requests
from selenium import webdriver
import demjson
from lxml import etree
import time


def get_excel(bookname_list):
    "将excel表中的数据加载至列表"

    
    workbook = load_workbook(filename='./douban/files/1.xlsx')
    sheet = workbook['Sheet3']
    for cell in sheet['B']:
        bookname_list.append(cell.value)
    return bookname_list





def get_subject_id(wd, bookname):
    "搜索关键词，获取subject_id"


    url = 'https://search.douban.com/book/subject_search?search_text=' + str(bookname)
    wd.get(url)

    element = wd.find_element_by_xpath("//div[starts-with(@class,'sc-bZQynM')]/div[@class='item-root']/a")
    subject_url =  element.get_attribute('href')

    return subject_url


def getdata(subject_url,book_data):
    "此函数用来接收并处理HTML文件"



    #接收HTML文件
    diy_headers = {
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36',
    }

    result = requests.get(subject_url,headers = diy_headers)
    soup_html = result.text
    soup = BeautifulSoup(soup_html, 'html.parser')

    #书名
    selector=etree.HTML(str(soup_html))
    datasoup = selector.xpath("//body/div[@id='wrapper']/h1/span")
    book_data['书名'] = datasoup[0].text

    
    #处理其他信息
    info_html = str(soup.find_all(id="info"))

    #处理简单信息
    selector=etree.HTML(info_html)
    datasoup = selector.xpath('//span')
    for item in datasoup:
        if item.text == '出版社:':
            book_data['出版社'] = item.tail.lstrip()
        elif item.text == '原作名:':
            book_data['原作名'] = item.tail.lstrip()
        elif item.text == '出版年:':
            book_data['出版年'] = item.tail.lstrip()
        elif item.text == '页数:':
            book_data['页数'] = item.tail.lstrip()
        elif item.text == '定价:':
            book_data['定价'] = item.tail.lstrip()
        elif item.text == '装帧:':
            book_data['装帧'] = item.tail.lstrip()
        elif item.text == 'ISBN:':
            book_data['ISBN'] = item.tail.lstrip()
        elif item.text == '副标题:':
            book_data['副标题'] = item.tail.lstrip()
        else:
            pass
    
    #处理作者
    author_form = re.compile("作者.*?<.?br.?>",flags=re.S)
    author_list = str(re.findall(author_form,info_html))
    author_select=etree.HTML(author_list)
    author_datasoup = author_select.xpath('//a')
    author_str = ''
    if len(author_datasoup)>1:
        for item in author_datasoup:
            author_str = author_str + item.text.replace('\\n','').replace(' ','') + '/'
        author_str = author_str[:-1]
    else:
        for item in author_datasoup:
            author_str = item.text.replace('\\n','').replace(' ','')
    book_data['作者'] = author_str

    #处理译者
    translate_form = re.compile("译者.*?<.?br.?>",flags=re.S)
    translate_list = str(re.findall(translate_form,info_html))
    translate_select=etree.HTML(translate_list)
    translate_datasoup = translate_select.xpath('//a')
    translate_str = ''
    if len(translate_datasoup)>1:
        for item in translate_datasoup:
            translate_str = translate_str + item.text.replace('\\n','').replace(' ','') + '/'
        translate_str = translate_str[:-1]
    else:
        for item in translate_datasoup:
            translate_str = item.text.replace('\\n','').replace(' ','')
    book_data['译者'] = translate_str

    #处理出品方
    producer_form = re.compile("出品方.*?<.?br.?>",flags=re.S)
    producer_list = str(re.findall(producer_form,info_html))
    producer_select=etree.HTML(producer_list)
    producer_datasoup = producer_select.xpath('//a')
    producer_str = ''
    if len(producer_datasoup)>1:
        for item in producer_datasoup:
            producer_str = producer_str + item.text.replace('\\n','').replace(' ','') + '/'
        producer_str = producer_str[:-1]
    else:
        for item in producer_datasoup:
            producer_str = item.text.replace('\\n','').replace(' ','')
    book_data['出品方'] = producer_str

    #处理丛书
    series_form = re.compile("丛书.*?<.?br.?>",flags=re.S)
    series_list = str(re.findall(series_form,info_html))
    series_select=etree.HTML(series_list)
    series_datasoup = series_select.xpath('//a')
    series_str = ''
    if len(series_datasoup)>1:
        for item in series_datasoup:
            series_str = series_str + item.text.replace('\\n','').replace(' ','') + '/'
        series_str = series_str[:-1]
    else:
        for item in series_datasoup:
            series_str = item.text.replace('\\n','').replace(' ','')
    book_data['丛书'] = series_str

    #处理简介
    intro_html = str(soup.find_all(id='link-report'))
    selector=etree.HTML(intro_html)
    datasoup = selector.xpath('//div/div/p')
    intro_str = ''
    for item in datasoup:
        intro_str = intro_str + str(item.text) + '\n'
    book_data['简介'] = intro_str[:-1]

    #处理图片链接并按照/img/ISBN_书名.jpg的形式命名保存
    img_html = str(soup.find_all(class_='nbg'))
    img_link_form = re.compile(r"https.*?l/public/.*?.jpg")
    img_link = re.findall(img_link_form,img_html)[0]
    book_data['封面'] = img_link
    saveimg(img_link,book_data)


    return


def savedata(book_data):
    "保存到excel"

    workbook = load_workbook(filename='./douban/files/database.xlsx')
    sheet = workbook['书单']
    sheet.append(list(book_data.values()))


    workbook.save(filename='./douban/files/database.xlsx')

def saveimg(url, book_data):
    "保存图片"

    diy_headers = {
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36',
    }
    picture = requests.get(url,headers = diy_headers,stream = True)
    print(picture.status_code) # 返回状态码
    if picture.status_code == 200:
        filepath = './douban/files/img/' + book_data['ISBN'] + '_' + book_data['书名'] + '.jpg'
        open(filepath, 'wb').write(picture.content) # 将内容写入图片
        print("done")
    del picture
    return


def main():
    #建立一个包含书名的列表bookname_list
    bookname_list = []
    bookname_list = get_excel(bookname_list)



    
    wd = webdriver.Chrome("./douban/chromedriver.exe")
    for bookname in bookname_list:

        #定义书本字典
        book_data = {
            "ISBN" : "",
            "书名" : "",
            "作者" : "",
            "出版社" : "",
            "出品方" : "",
            "副标题" : "",
            "原作名" : "",
            "译者" : "",
            "出版年" : "",
            "页数" : "",
            "定价" : "",
            "装帧" : "",
            "丛书" : "",
            "国籍" : "",
            "简介" : "",
            "封面" : ""
        }

        subject_url =  get_subject_id(wd, bookname)
        time.sleep(10)
        getdata(subject_url, book_data)
        time.sleep(10)
        savedata(book_data) 
    wd.quit()
    



if __name__ == "__main__":
    main()