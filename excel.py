#-*- encoding=utf-8 -*-

#excel模块
from openpyxl import load_workbook          #读取excel文件
from openpyxl import Workbook               #创建excel文件
from openpyxl.styles import Font            #修改字体样式
from openpyxl.styles import Alignment       #设置对齐样式
from openpyxl.styles import Side,Border     #边框样式
from openpyxl.styles import PatternFill     #纯色填充
from openpyxl.styles import GradientFill    #渐变填充

#打开excel表格并获取sheet名称
# workbook = load_workbook(filename='./files/test.xlsx', data_only=True)
# data_only=True 表示计算出表格中公式的结果
workbook = load_workbook(filename='./files/test.xlsx')
print(workbook.sheetnames)
# 通过sheet名称获取表格
sheet = workbook['书单']
#也可以通过workbook.active来获取默认的sheet
# sheet=workbook.active
'''
# 获取表格的尺寸大小
print(sheet.dimensions)

#获取表格内某个格子的数据
# .row          行数
# .cloumn       列数
# .coordinate   坐标
cell = sheet['G1']
print(cell.value, cell.row, cell.column, cell.coordinate)


#获取当前sheet的合并单元格
for mergeds in sheet.merged_cells.ranges:
    # print(mergeds.max_row, mergeds.min_row, mergeds.max_col, mergeds.min_col)
    print(mergeds)

#合并单元格的内容存放在最左上角的那个格子里
cell = sheet['G1']
print(cell.value)


# 使用坐标定位格子
cell = sheet.cell(row=6,column=8)
print(cell.coordinate)

#获取一系列格子
cells = sheet['A']
cells = sheet['A:C']
cells = sheet[5]
cells = sheet[5:6]
#指定行和列的范围，按行获取
cells = sheet.iter_rows(min_row=2,max_row=3,min_col=2,max_col=3)
#指定行和列的范围，按列获取
cells = sheet.iter_cols(min_row=2,max_row=3,min_col=2,max_col=3)

for item in sheet["A3:B10"]:
    for cell in item:
        print(cell.value)

#迭代表格的所有行
for row in sheet.rows:
    for cell in row:
        if cell.value:
            print(cell.value)
#迭代表格的所有列
# sheet.columns


#向某个格子写入内容并保存
sheet = workbook['Sheet1']
sheet['A1'] = '世界'
# 另一种写法
# cell = sheet['A1']
# cell.value = '世界'


# 用python列表数据插入一行
data = [
    ['a' , 1 ],
    ['b' , 2 ],
    ['c' , 3 ],
    ['d' , 4 ],
]

for row in data:
    sheet.append(row)

#插入公式
sheet['C1'] = '=SUM(B3:B9)'

#插入空白列
#在idx列左边插入一列
#.insert_cols(idx=数字编号, amount=要插入的列数)
sheet.insert_cols(idx=2, amount=3)

#插入空白行
#在idx行上面插入一行
#insert_rows(idx=数字编号, amount=要插入的行数)
sheet.insert_rows(idx=2, amount=3)

#删除列
#.delete_cols(idx=数字编号, amount=要删除的列数)
#从idx这一列开始，包括idx这一列
sheet.delete_cols(idx=2, amount=3)

#删除行
#.delete_rows(idx=数字编号, amount=要删除的行数)
#从idx这一行开始，包括idx这一行
sheet.delete_rows(idx=2, amount=3)

#移动格子
# 正整数为向下或向右
# 负整数为向左或向上
sheet.move_range('A6:B11', rows=2, cols=2)

#创建新的sheet
workbook.create_sheet('这是一个新建sheet')
print(workbook.sheetnames)

#复制一个sheet
sheet = workbook['这是一个sheet']
workbook.copy_worksheet(sheet)
print(workbook.sheetnames)

#删除sheet
sheet = workbook['这是一个sheet Copy']
workbook.remove(sheet)
print(workbook.sheetnames)

#修改sheet名称
sheet.title = '更改后的名称'
print(workbook.sheetnames)
sheet.title = 'Sheet1'
print(workbook.sheetnames)

#冻结窗格
sheet.freeze_panes = 'B2'

#添加筛选
sheet.auto_filter.ref = sheet.dimensions

#创建新的excel文件
#Workbook注意大小写
workbook = Workbook()
sheet = workbook.active
sheet.title = '表格1'

workbook.save(filename='./files/newtest.xlsx')

'''

#修改字体样式
# name      字体名称
# size      字体大小
# bold      是否加粗
# italic    是否斜体
# color     字体颜色
font = Font(name='微软雅黑', size=12, bold=True, italic=True, color='bf242a')
cell = sheet['A2']
cell.font = font

#获取字体样式
cell = sheet['B4']
font = cell.font
print(font.name, font.size, font.bold, font.italic)

#设置对齐样式
# horizontal    水平对齐
#           可选项 distributed      分散对齐
#           可选项 justify          两端对齐
#           可选项 center           居中
#           可选项 left             靠左(缩进)
#           可选项 fill             填充
#           可选项 centerContinuous 跨列居中
#           可选项 right            靠右(缩进)
#           可选项 general          常规

# vertical      垂直对齐
#           可选项 bottom       靠下
#           可选项 distributed  分散对齐
#           可选项 justify      两端对齐
#           可选项 center       居中
#           可选项 top          靠上

# text_rotation 旋转角度
# wrap_text     是否自动换行
cell = sheet['B2']
alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True)
cell.alignment = alignment

#设置边框样式
# Side(style=边线样式, color=边线颜色)
# Border(left=左边线样式, right=右边线样式, top=上边线样式, bottom=下边线样式)
cell = sheet['B3']
side = Side(style='thin', color='bf242a')
border = Border(left=side, right=side, top=side, bottom=side)
cell.border = border

#设置填充样式
# PatternFill(fill_type=填充样式, fgColor='填充颜色')
# GradientFill(stop=('渐变颜色1', '渐变颜色2', '渐变颜色3', ......))
cell = sheet['B5']
pattern_fill = PatternFill(fill_type='solid', fgColor='bf242a')
cell.fill = pattern_fill
cell = sheet['B6']
gradient_fill = GradientFill(stop=('bf242a','47585c','cd5e3c'))
cell.fill = gradient_fill

#设置行高与列宽
sheet.row_dimensions[1].height = 30
sheet.column_dimensions['B'].width = 20

#合并单元格
#sheet.merge_cells('C12:E16')
sheet.merge_cells(start_row=12, start_column=3, end_row=16, end_column=5)

#取消合并单元格
#sheet.merge_cells('C12:E16')
sheet.unmerge_cells(start_row=12, start_column=3, end_row=16, end_column=5)


workbook.save(filename='./files/test.xlsx')